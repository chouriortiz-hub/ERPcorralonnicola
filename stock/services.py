"""
Servicios de STOCK: orquestación de las funciones "grandes" del módulo que
no son un simple CRUD (boletas con sus líneas, corrección/anulación
retroactiva de boletas, importación/exportación a Excel). Separado de
views.py siguiendo el mismo patrón que `ventas/services.py`, para que esta
lógica se pueda reutilizar o testear sin pasar por una request HTTP.
"""
import io
import unicodedata
from decimal import Decimal, InvalidOperation

from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import F, Sum
from openpyxl import Workbook, load_workbook

from .models import Boleta, BoletaItem, MovimientoStock, Nota, NotaProducto, Producto, registrar_movimiento


# ---------------------------------------------------------------------------
# Boletas: alta, corrección retroactiva y anulación
# ---------------------------------------------------------------------------

def parsear_items_boleta(items):
    """
    Valida y normaliza la lista de líneas de una boleta (ya deserializada de
    JSON, mismo formato que llega del "carrito" de líneas en el formulario):
        [{"producto_id": 1, "cantidad": "5", "nota": "...",
          "texto_original": "...", "confianza": 87.5}, ...]
    -> [{"producto": <Producto>, "cantidad": Decimal, "nota": str,
         "texto_original": str, "confianza": Decimal|None}, ...]
    """
    if not isinstance(items, list) or not items:
        raise ValidationError('La boleta no tiene líneas de productos cargadas.')

    try:
        producto_ids = [int(item['producto_id']) for item in items]
    except (KeyError, TypeError, ValueError):
        raise ValidationError('La boleta tiene un producto inválido.')

    productos = Producto.objects.in_bulk(producto_ids)

    lineas = []
    for item in items:
        producto_id = int(item['producto_id'])
        producto = productos.get(producto_id)
        if producto is None or not producto.activo:
            raise ValidationError(f'El producto seleccionado (id {producto_id}) no existe o no está activo.')

        try:
            cantidad = Decimal(str(item.get('cantidad', '0')))
        except InvalidOperation:
            raise ValidationError(f'Cantidad inválida para "{producto.nombre}".')
        if cantidad <= 0:
            raise ValidationError(f'La cantidad de "{producto.nombre}" debe ser mayor a cero.')

        confianza = item.get('confianza')
        try:
            confianza = Decimal(str(confianza)) if confianza not in (None, '') else None
        except InvalidOperation:
            confianza = None

        lineas.append({
            'producto': producto,
            'cantidad': cantidad,
            'nota': (item.get('nota') or '')[:200],
            'texto_original': (item.get('texto_original') or '')[:250],
            'confianza': confianza,
        })
    return lineas


@transaction.atomic
def crear_boleta(boleta, items, usuario):
    """
    Guarda la cabecera de la Boleta (ya validada por BoletaForm, todavía sin
    persistir), crea sus líneas y aplica de una vez los movimientos de stock
    correspondientes (ENTRADA o SALIDA según `boleta.tipo`). Si alguna línea
    deja stock negativo, `registrar_movimiento` aborta y toda la operación
    se revierte (incluida la creación de la boleta).
    """
    boleta.usuario = usuario
    boleta.estado = Boleta.CONFIRMADA
    boleta.save()

    for linea in items:
        BoletaItem.objects.create(
            boleta=boleta,
            producto=linea['producto'],
            cantidad=linea['cantidad'],
            nota=linea['nota'],
            texto_original=linea['texto_original'],
            confianza=linea['confianza'],
        )
        registrar_movimiento(
            producto=linea['producto'],
            cantidad=linea['cantidad'],
            tipo=boleta.tipo,
            origen=MovimientoStock.ORIGEN_BOLETA,
            usuario=usuario,
            referencia_id=boleta.pk,
            motivo=f'Boleta #{boleta.numero}',
        )
    return boleta


@transaction.atomic
def ajustar_boleta(boleta, nuevos_items, usuario):
    """
    Corrección retroactiva de una boleta ya confirmada: revierte los
    movimientos de las líneas actuales y aplica los de las líneas nuevas,
    todo en una única transacción. Si la reversión o la nueva aplicación
    dejarían stock negativo en cualquier paso, `registrar_movimiento` lanza
    ValidationError y la transacción completa se descarta — la boleta queda
    exactamente como estaba antes de intentar el ajuste.
    """
    if boleta.estado != Boleta.CONFIRMADA:
        raise ValidationError('Solo se pueden ajustar boletas confirmadas.')

    tipo_reversion = MovimientoStock.SALIDA if boleta.tipo == Boleta.ENTRADA else MovimientoStock.ENTRADA
    for item in boleta.items.select_related('producto'):
        registrar_movimiento(
            producto=item.producto,
            cantidad=item.cantidad,
            tipo=tipo_reversion,
            origen=MovimientoStock.ORIGEN_AJUSTE_BOLETA,
            usuario=usuario,
            referencia_id=boleta.pk,
            motivo=f'Corrección boleta #{boleta.numero} (reversión de línea anterior)',
        )
    boleta.items.all().delete()

    for linea in nuevos_items:
        BoletaItem.objects.create(
            boleta=boleta,
            producto=linea['producto'],
            cantidad=linea['cantidad'],
            nota=linea['nota'],
            texto_original=linea['texto_original'],
            confianza=linea['confianza'],
        )
        registrar_movimiento(
            producto=linea['producto'],
            cantidad=linea['cantidad'],
            tipo=boleta.tipo,
            origen=MovimientoStock.ORIGEN_AJUSTE_BOLETA,
            usuario=usuario,
            referencia_id=boleta.pk,
            motivo=f'Corrección boleta #{boleta.numero} (línea corregida)',
        )
    return boleta


@transaction.atomic
def anular_boleta(boleta, usuario, motivo=''):
    """
    Anula una boleta confirmada revirtiendo sus movimientos (nunca se borra
    ni la boleta ni sus líneas, para conservar la trazabilidad completa de
    qué pasó). Si la reversión dejaría stock negativo, se aborta.
    """
    if boleta.estado != Boleta.CONFIRMADA:
        raise ValidationError('Solo se pueden anular boletas confirmadas.')

    tipo_reversion = MovimientoStock.SALIDA if boleta.tipo == Boleta.ENTRADA else MovimientoStock.ENTRADA
    for item in boleta.items.select_related('producto'):
        registrar_movimiento(
            producto=item.producto,
            cantidad=item.cantidad,
            tipo=tipo_reversion,
            origen=MovimientoStock.ORIGEN_AJUSTE_BOLETA,
            usuario=usuario,
            referencia_id=boleta.pk,
            motivo=f'Anulación boleta #{boleta.numero}' + (f' — {motivo}' if motivo else ''),
        )
    boleta.estado = Boleta.ANULADA
    boleta.save(update_fields=['estado', 'actualizado'])
    return boleta


# ---------------------------------------------------------------------------
# Notas: bitácora con productos agendados (snapshot) y movimientos puntuales
# ---------------------------------------------------------------------------

@transaction.atomic
def crear_nota(nota, usuario, producto_ids, movimiento_ids):
    """
    Guarda la Nota (cabecera ya validada por NotaForm, todavía sin
    persistir) y agenda sus referencias:
    - Por cada producto seleccionado, congela un `NotaProducto` con el
      stock/código/nombre de ESE momento — no se vuelve a tocar aunque el
      producto cambie después (ver docstring del modelo).
    - Los movimientos seleccionados se linkean directo por M2M: un
      MovimientoStock ya es inmutable por diseño, no necesita snapshot.
    """
    if not isinstance(producto_ids, list) or not isinstance(movimiento_ids, list):
        raise ValidationError('La selección de productos o movimientos es inválida.')

    try:
        producto_ids = [int(i) for i in producto_ids]
        movimiento_ids = [int(i) for i in movimiento_ids]
    except (TypeError, ValueError):
        raise ValidationError('La selección de productos o movimientos tiene un id inválido.')

    productos = Producto.objects.in_bulk(producto_ids)
    faltantes = set(producto_ids) - set(productos.keys())
    if faltantes:
        raise ValidationError(f'Hay producto(s) seleccionados que ya no existen (id {sorted(faltantes)}).')

    movimientos = MovimientoStock.objects.filter(pk__in=movimiento_ids)
    if movimientos.count() != len(set(movimiento_ids)):
        raise ValidationError('Hay movimiento(s) seleccionados que ya no existen.')

    nota.usuario = usuario
    nota.save()

    for pid in producto_ids:
        producto = productos[pid]
        NotaProducto.objects.create(
            nota=nota, producto=producto,
            codigo_registrado=producto.codigo, nombre_registrado=producto.nombre,
            stock_registrado=producto.stock_actual, unidad_medida_registrada=producto.unidad_medida,
        )

    if movimiento_ids:
        nota.movimientos.set(movimientos)

    return nota


# ---------------------------------------------------------------------------
# Historial de movimientos: contadores del período filtrado
# ---------------------------------------------------------------------------

def resumen_movimientos(movimientos):
    """
    Métricas agregadas sobre el queryset de MovimientoStock YA FILTRADO
    (mismo filtro de q/tipo/período que se está mostrando en la grilla del
    historial), para las tarjetas de resumen arriba de la tabla.
    """
    entradas = movimientos.filter(tipo=MovimientoStock.ENTRADA)
    salidas = movimientos.filter(tipo=MovimientoStock.SALIDA)

    entradas_unidades = entradas.aggregate(total=Sum('cantidad'))['total'] or Decimal('0')
    salidas_unidades = salidas.aggregate(total=Sum('cantidad'))['total'] or Decimal('0')

    boletas_detectadas = (
        movimientos.filter(origen=MovimientoStock.ORIGEN_BOLETA)
        .values('referencia_id').distinct().count()
    )

    return {
        'total': movimientos.count(),
        'entradas_count': entradas.count(),
        'entradas_unidades': entradas_unidades,
        'salidas_count': salidas.count(),
        'salidas_unidades': salidas_unidades,
        'balance_unidades': entradas_unidades - salidas_unidades,
        'productos_tocados': movimientos.values('producto_id').distinct().count(),
        'boletas_detectadas': boletas_detectadas,
    }


# ---------------------------------------------------------------------------
# Dashboard / alertas / valorización
# ---------------------------------------------------------------------------

def resumen_stock():
    """Métricas agregadas para el dashboard de stock (alertas + valorización)."""
    productos = Producto.objects.filter(activo=True)
    con_control = productos.filter(descuenta_stock=True)

    sin_stock = con_control.filter(stock_actual__lte=0)
    bajo_minimo = con_control.filter(stock_actual__gt=0, stock_actual__lte=F('stock_minimo'), stock_minimo__gt=0)
    ok = con_control.exclude(pk__in=sin_stock.values('pk')).exclude(pk__in=bajo_minimo.values('pk'))

    valor_total = sum((p.valor_stock for p in productos), start=Decimal('0'))

    return {
        'total_productos': productos.count(),
        'con_control': con_control.count(),
        'ok': ok.count(),
        'bajo_minimo': bajo_minimo.count(),
        'sin_stock': sin_stock.count(),
        'valor_total': valor_total,
        'productos_alerta': (bajo_minimo | sin_stock).distinct().order_by('stock_actual'),
    }


# ---------------------------------------------------------------------------
# Excel: importación masiva de stock inicial y exportación general
# ---------------------------------------------------------------------------

def _normalizar_encabezado(valor):
    if valor is None:
        return ''
    texto = str(valor).strip().upper()
    texto = unicodedata.normalize('NFKD', texto).encode('ascii', 'ignore').decode('ascii')
    return texto


@transaction.atomic
def importar_stock_excel(archivo, usuario):
    """
    Lee un .xlsx con al menos una columna de código de producto y una de
    stock, detectando la fila de encabezado automáticamente entre las
    primeras filas de la primera hoja (o de una hoja cuyo nombre contenga
    "STOCK"). Por cada código que coincida con un producto existente,
    registra un AJUSTE por la diferencia entre el valor de la planilla y el
    stock actual (no crea productos nuevos ni pisa el campo directamente,
    para que quede rastro en el historial de movimientos).
    """
    try:
        libro = load_workbook(archivo, data_only=True)
    except Exception:
        raise ValidationError('No se pudo leer el archivo. Verificá que sea un Excel (.xlsx) válido.')

    hoja = None
    for nombre in libro.sheetnames:
        if 'STOCK' in nombre.upper():
            hoja = libro[nombre]
            break
    if hoja is None:
        hoja = libro[libro.sheetnames[0]]

    col_codigo = col_stock = None
    fila_encabezado = None
    for fila in hoja.iter_rows(min_row=1, max_row=min(10, hoja.max_row)):
        valores = [_normalizar_encabezado(c.value) for c in fila]
        try:
            idx_codigo = next(i for i, v in enumerate(valores) if 'CODIGO' in v)
        except StopIteration:
            continue
        try:
            idx_stock = next(i for i, v in enumerate(valores) if v in ('STOCK', 'CANTIDAD', 'STOCK INICIAL', 'STOCK ACTUAL'))
        except StopIteration:
            continue
        col_codigo, col_stock, fila_encabezado = idx_codigo, idx_stock, fila[0].row
        break

    if fila_encabezado is None:
        raise ValidationError(
            'No se encontró una fila de encabezado con columnas de "Código" y "Stock" '
            'en las primeras 10 filas de la planilla.'
        )

    productos_por_codigo = {p.codigo.upper(): p for p in Producto.objects.all()}
    actualizados, sin_cambios, no_encontrados = 0, 0, []

    for fila in hoja.iter_rows(min_row=fila_encabezado + 1):
        codigo_val = fila[col_codigo].value
        if codigo_val in (None, ''):
            continue
        codigo = str(codigo_val).strip().upper()
        producto = productos_por_codigo.get(codigo)
        if producto is None:
            no_encontrados.append(codigo)
            continue

        try:
            nuevo_stock = Decimal(str(fila[col_stock].value or 0))
        except InvalidOperation:
            no_encontrados.append(codigo)
            continue

        delta = nuevo_stock - producto.stock_actual
        if delta == 0:
            sin_cambios += 1
            continue

        registrar_movimiento(
            producto=producto,
            cantidad=abs(delta),
            tipo=MovimientoStock.ENTRADA if delta > 0 else MovimientoStock.SALIDA,
            origen=MovimientoStock.ORIGEN_IMPORTACION,
            usuario=usuario,
            motivo='Importación masiva desde Excel',
        )
        actualizados += 1

    return {'actualizados': actualizados, 'sin_cambios': sin_cambios, 'no_encontrados': no_encontrados}


def exportar_stock_excel():
    """Arma el .xlsx multi-hoja (Stock, Entradas, Salidas, Alertas, Boletas) y
    devuelve los bytes listos para servir como descarga."""
    wb = Workbook()

    ws_stock = wb.active
    ws_stock.title = 'STOCK'
    ws_stock.append(['Código', 'Nombre', 'Categoría', 'Unidad', 'Stock actual', 'Stock mínimo', 'PMP', 'Precio venta', 'Valor total', 'Descuenta stock', 'Activo'])
    valor_total_general = Decimal('0')
    for p in Producto.objects.select_related('categoria').order_by('nombre'):
        valor = p.valor_stock
        valor_total_general += valor
        ws_stock.append([
            p.codigo, p.nombre, p.categoria.nombre, p.get_unidad_medida_display(),
            float(p.stock_actual), float(p.stock_minimo), float(p.pmp), float(p.precio_venta), float(valor),
            'Sí' if p.descuenta_stock else 'No', 'Sí' if p.activo else 'No',
        ])
    ws_stock.append([])
    ws_stock.append(['', '', '', '', '', '', '', 'VALOR TOTAL', float(valor_total_general)])

    def hoja_movimientos(nombre, tipo):
        ws = wb.create_sheet(nombre)
        ws.append(['Fecha', 'Código', 'Producto', 'Cantidad', 'Stock resultante', 'Origen', 'Usuario', 'Motivo'])
        qs = MovimientoStock.objects.select_related('producto', 'usuario').filter(tipo=tipo).order_by('-fecha')
        for m in qs:
            ws.append([
                m.fecha.strftime('%d/%m/%Y %H:%M'), m.producto.codigo, m.producto.nombre,
                float(m.cantidad), float(m.stock_resultante), m.get_origen_display(),
                str(m.usuario), m.motivo,
            ])

    hoja_movimientos('ENTRADAS', MovimientoStock.ENTRADA)
    hoja_movimientos('SALIDAS', MovimientoStock.SALIDA)

    ws_alertas = wb.create_sheet('ALERTAS')
    ws_alertas.append(['Código', 'Nombre', 'Stock actual', 'Stock mínimo'])
    for p in resumen_stock()['productos_alerta']:
        ws_alertas.append([p.codigo, p.nombre, float(p.stock_actual), float(p.stock_minimo)])

    ws_boletas = wb.create_sheet('BOLETAS')
    ws_boletas.append(['Número', 'Tipo', 'Fecha', 'Responsable', 'Monto', 'Estado', 'Items'])
    for b in Boleta.objects.prefetch_related('items__producto').order_by('-fecha'):
        items_txt = '; '.join(f'{i.producto.codigo} x {i.cantidad}' for i in b.items.all())
        ws_boletas.append([b.numero, b.get_tipo_display(), b.fecha.strftime('%d/%m/%Y'), b.responsable, float(b.monto), b.get_estado_display(), items_txt])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
